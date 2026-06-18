from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from django.contrib.auth.models import User
from openpyxl import load_workbook

from .models import Report, ReportData
from .serializers import ReportSerializer, ReportUploadSerializer, ReportDataSerializer, UserSerializer, RegisterSerializer
from .permissions import IsAdminUser


class ReportDataPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 200


class ReportPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100


class ReportViewSet(viewsets.ModelViewSet):
    queryset = Report.objects.all()
    serializer_class = ReportSerializer
    pagination_class = ReportPagination

    def get_queryset(self):
        return Report.objects.select_related('uploaded_by')

    def get_permissions(self):
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdminUser()]
        return [IsAuthenticated()]

    def get_serializer_class(self):
        if self.action == 'create':
            return ReportUploadSerializer
        return ReportSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        file = serializer.validated_data['file']
        title = serializer.validated_data['title']

        report = Report.objects.create(
            title=title,
            file=file,
            uploaded_by=request.user
        )

        try:
            file.seek(0)
            workbook = load_workbook(file, read_only=True, data_only=True)
            worksheet = workbook.active
            non_empty_columns = set()
            header_candidates = []
            non_empty_row_count = 0

            for raw_row in worksheet.iter_rows(values_only=True):
                if not raw_row or not self._row_has_values(raw_row):
                    continue

                non_empty_row_count += 1
                for col_idx, value in enumerate(raw_row):
                    if self._has_value(value):
                        non_empty_columns.add(col_idx)

                if len(header_candidates) < 10:
                    header_candidates.append(tuple(raw_row))

            workbook.close()

            if non_empty_row_count == 0:
                raise ValueError("Excel file is empty")

            non_empty_columns = sorted(non_empty_columns)
            pruned_candidates = [
                self._prune_row(row, non_empty_columns)
                for row in header_candidates
            ]

            header_idx = 0
            num_cols = len(non_empty_columns)
            for idx, row in enumerate(pruned_candidates):
                non_empty = sum(1 for val in row if self._has_value(val))
                if non_empty >= num_cols * 0.3:
                    header_idx = idx
                    break

            header_row = pruned_candidates[header_idx]
            headers = []
            used_names = set()
            for i in range(num_cols):
                val = header_row[i] if i < len(header_row) else None
                if val is None or str(val).strip() == '':
                    name = f'col_{i}'
                else:
                    name = str(val).strip()
                if name in used_names:
                    name = f'{name}_{i}'
                used_names.add(name)
                headers.append(name)

            file.seek(0)
            workbook = load_workbook(file, read_only=True, data_only=True)
            worksheet = workbook.active

            report_data_objects = []
            row_num = 1
            non_empty_seen = 0
            for raw_row in worksheet.iter_rows(values_only=True):
                if not raw_row or not self._row_has_values(raw_row):
                    continue

                if non_empty_seen <= header_idx:
                    non_empty_seen += 1
                    continue

                non_empty_seen += 1
                row = self._prune_row(raw_row, non_empty_columns)
                row_data = {}
                for i, header in enumerate(headers):
                    value = row[i] if i < len(row) else None
                    if value is None or str(value).strip() == '':
                        row_data[header] = None
                    else:
                        try:
                            num_val = float(value)
                            row_data[header] = num_val if num_val != int(num_val) else int(num_val)
                        except (ValueError, TypeError):
                            row_data[header] = str(value)

                report_data_objects.append(
                    ReportData(
                        report=report,
                        row_number=row_num,
                        data=row_data
                    )
                )
                row_num += 1

                if len(report_data_objects) >= 1000:
                    ReportData.objects.bulk_create(report_data_objects, batch_size=1000)
                    report_data_objects = []

            workbook.close()

            if report_data_objects:
                ReportData.objects.bulk_create(report_data_objects, batch_size=1000)

            # Save headers to preserve column order
            report.headers = headers
            report.row_count = row_num - 1
            report.save(update_fields=['headers', 'row_count'])

        except Exception as e:
            report.delete()
            return Response(
                {'error': f'Failed to process Excel file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response(
            ReportSerializer(report).data,
            status=status.HTTP_201_CREATED
        )

    @staticmethod
    def _has_value(value):
        return value is not None and str(value).strip() != ''

    @classmethod
    def _row_has_values(cls, row):
        return any(cls._has_value(value) for value in row)

    @staticmethod
    def _prune_row(row, columns):
        return tuple(row[col_idx] if col_idx < len(row) else None for col_idx in columns)

    @action(detail=True, methods=['get'])
    def data(self, request, pk=None):
        report = self.get_object()
        paginator = ReportDataPagination()
        data_rows = report.data_rows.only('id', 'row_number', 'data').order_by('row_number')
        page = paginator.paginate_queryset(data_rows, request)

        if page is not None:
            serializer = ReportDataSerializer(page, many=True)
            response = paginator.get_paginated_response(serializer.data)
            # Use stored headers to preserve column order
            headers = report.headers if report.headers else []
            # Fallback for old reports without stored headers
            if not headers and data_rows.exists():
                first_row = data_rows.first()
                if first_row and first_row.data:
                    headers = list(first_row.data.keys())
            response.data['headers'] = headers
            return response

        serializer = ReportDataSerializer(data_rows, many=True)
        return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def current_user(request):
    serializer = UserSerializer(request.user)
    return Response(serializer.data)


@api_view(['POST'])
@permission_classes([IsAdminUser])
def register(request):
    serializer = RegisterSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.save()
        return Response({
            'message': 'Registration successful',
            'user': UserSerializer(user).data
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def list_users(request):
    users = User.objects.all().order_by('-date_joined')
    serializer = UserSerializer(users, many=True)
    return Response(serializer.data)


@api_view(['DELETE'])
@permission_classes([IsAdminUser])
def delete_user(request, user_id):
    try:
        user = User.objects.get(id=user_id)
        if user == request.user:
            return Response(
                {'error': 'Cannot delete yourself'},
                status=status.HTTP_400_BAD_REQUEST
            )
        username = user.username
        user.delete()
        return Response({'message': f'User {username} deleted successfully'})
    except User.DoesNotExist:
        return Response(
            {'error': 'User not found'},
            status=status.HTTP_404_NOT_FOUND
        )
